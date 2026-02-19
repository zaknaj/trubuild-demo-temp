"""
Excel BOQ extraction module using openpyxl.

This module handles the extraction of BOQ data from Excel files (.xlsx, .xls).
Excel BOQs are often more structured than PDFs, with clear columns for
item codes, descriptions, quantities, units, and rates.

Supported formats:
- .xlsx (Excel 2007+)
- .xls (Legacy Excel) - requires xlrd

Usage:
    from parse_excel import ExcelParser, extract_from_excel
    
    # Simple extraction
    text = extract_from_excel("boq.xlsx")
    
    # Detailed extraction with structure
    parser = ExcelParser("boq.xlsx")
    result = parser.extract()
"""

from __future__ import annotations

import re
from pathlib import Path
from dataclasses import dataclass, field
from typing import Any
from enum import Enum, auto

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import Cell
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ColumnType(Enum):
    """Types of columns commonly found in BOQ spreadsheets."""
    ITEM_ID = auto()
    DESCRIPTION = auto()
    QUANTITY = auto()
    UNIT = auto()
    RATE = auto()
    AMOUNT = auto()
    UNKNOWN = auto()


@dataclass
class ExcelColumn:
    """
    Represents a detected column in the BOQ spreadsheet.

    Attributes:
        index: 0-based column index
        letter: Excel column letter (A, B, C, ...)
        header: Header text
        column_type: Detected column type
        confidence: Detection confidence (0-1)
    """
    index: int
    letter: str
    header: str
    column_type: ColumnType
    confidence: float = 0.0


@dataclass
class ExcelRow:
    """
    Represents a row extracted from the spreadsheet.

    Attributes:
        row_number: 1-based row number
        cells: Dictionary mapping column letter to cell value
        is_header: Whether this row appears to be a header
        is_division: Whether this row appears to be a division header
        is_grouping: Whether this row appears to be a grouping header
        is_item: Whether this row appears to be a line item
    """
    row_number: int
    cells: dict[str, Any]
    is_header: bool = False
    is_division: bool = False
    is_grouping: bool = False
    is_item: bool = False
    raw_values: list[Any] = field(default_factory=list)


@dataclass
class ExcelSheet:
    """
    Represents a worksheet from the Excel file.

    Attributes:
        name: Sheet name
        rows: List of extracted rows
        columns: Detected column structure
        header_row: Row number of the header (if detected)
    """
    name: str
    rows: list[ExcelRow]
    columns: list[ExcelColumn]
    header_row: int | None = None


@dataclass
class ExcelExtractionResult:
    """
    Result of Excel extraction.

    Attributes:
        sheets: List of extracted sheets
        text: Combined text representation
        source_file: Path to source file
    """
    sheets: list[ExcelSheet]
    text: str
    source_file: str

    @property
    def primary_sheet(self) -> ExcelSheet | None:
        """Get the primary (first) sheet."""
        return self.sheets[0] if self.sheets else None


class ExcelParser:
    """
    Parser for extracting BOQ data from Excel files.

    Automatically detects column structure and extracts data
    in a format suitable for LLM processing.
    """

    # Common header patterns for column detection
    COLUMN_PATTERNS = {
        ColumnType.ITEM_ID: [
            r"item\s*(?:no|number|id|code|ref)?",
            r"(?:no|number|id|code|ref)\.?",
            r"s/?n",
            r"#",
        ],
        ColumnType.DESCRIPTION: [
            r"description",
            r"desc\.?",
            r"particulars",
            r"specification",
            r"work\s*item",
            r"scope",
        ],
        ColumnType.QUANTITY: [
            r"qty\.?",
            r"quantity",
            r"quan\.?",
        ],
        ColumnType.UNIT: [
            r"unit",
            r"uom",
            r"u/?m",
        ],
        ColumnType.RATE: [
            r"rate",
            r"unit\s*(?:price|rate|cost)",
            r"price",
        ],
        ColumnType.AMOUNT: [
            r"amount",
            r"total",
            r"value",
            r"sum",
        ],
    }

    def __init__(self, file_path: str | Path):
        """
        Initialize the Excel parser.

        Args:
            file_path: Path to the Excel file

        Raises:
            ImportError: If openpyxl is not installed
            FileNotFoundError: If file doesn't exist
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel parsing. "
                "Install with: pip install openpyxl"
            )

        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        self._workbook = None

    def extract(self) -> ExcelExtractionResult:
        """
        Extract BOQ data from the Excel file.

        Returns:
            ExcelExtractionResult with extracted data
        """
        self._workbook = openpyxl.load_workbook(
            self.file_path,
            data_only=True  # Get calculated values, not formulas
        )

        sheets = []
        all_text_parts = []

        for sheet_name in self._workbook.sheetnames:
            worksheet = self._workbook[sheet_name]
            sheet_result = self._extract_sheet(worksheet)
            sheets.append(sheet_result)

            # Build text representation
            sheet_text = self._sheet_to_text(sheet_result)
            if sheet_text.strip():
                all_text_parts.append(f"=== Sheet: {sheet_name} ===\n{sheet_text}")

        self._workbook.close()

        return ExcelExtractionResult(
            sheets=sheets,
            text="\n\n".join(all_text_parts),
            source_file=str(self.file_path)
        )

    def _extract_sheet(self, worksheet: Worksheet) -> ExcelSheet:
        """
        Extract data from a single worksheet.

        Args:
            worksheet: openpyxl Worksheet object

        Returns:
            ExcelSheet with extracted data
        """
        rows = []
        columns = []
        header_row = None

        # Find the header row and detect columns
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row)), start=1):
            row_values = [cell.value for cell in row]
            
            # Check if this looks like a header row
            if self._is_header_row(row_values):
                header_row = row_idx
                columns = self._detect_columns(row_values)
                break

        # Extract all rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1), start=1):
            row_values = [cell.value for cell in row]
            
            # Skip completely empty rows
            if all(v is None or str(v).strip() == "" for v in row_values):
                continue

            # Build cell dictionary
            cells = {}
            for col_idx, cell in enumerate(row):
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                cells[col_letter] = cell.value

            # Classify the row
            excel_row = ExcelRow(
                row_number=row_idx,
                cells=cells,
                raw_values=row_values,
                is_header=(row_idx == header_row),
            )

            # Detect row type
            if not excel_row.is_header:
                self._classify_row(excel_row, columns)

            rows.append(excel_row)

        return ExcelSheet(
            name=worksheet.title,
            rows=rows,
            columns=columns,
            header_row=header_row
        )

    def _is_header_row(self, row_values: list[Any]) -> bool:
        """
        Check if a row appears to be a header row.

        Args:
            row_values: List of cell values

        Returns:
            True if row looks like a header
        """
        if not row_values:
            return False

        # Count how many cells match known header patterns
        matches = 0
        non_empty = 0

        for value in row_values:
            if value is None or str(value).strip() == "":
                continue
            non_empty += 1
            value_str = str(value).lower().strip()

            for patterns in self.COLUMN_PATTERNS.values():
                for pattern in patterns:
                    if re.search(pattern, value_str, re.IGNORECASE):
                        matches += 1
                        break

        # Consider it a header if we have multiple matches
        return matches >= 2 and non_empty >= 3

    def _detect_columns(self, header_values: list[Any]) -> list[ExcelColumn]:
        """
        Detect column types from header row.

        Args:
            header_values: List of header cell values

        Returns:
            List of detected columns
        """
        columns = []

        for col_idx, value in enumerate(header_values):
            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
            header_text = str(value).strip() if value else ""

            # Detect column type
            col_type = ColumnType.UNKNOWN
            confidence = 0.0

            if header_text:
                header_lower = header_text.lower()
                for ctype, patterns in self.COLUMN_PATTERNS.items():
                    for pattern in patterns:
                        if re.search(pattern, header_lower, re.IGNORECASE):
                            col_type = ctype
                            confidence = 0.9
                            break
                    if col_type != ColumnType.UNKNOWN:
                        break

            columns.append(ExcelColumn(
                index=col_idx,
                letter=col_letter,
                header=header_text,
                column_type=col_type,
                confidence=confidence
            ))

        return columns

    def _classify_row(self, row: ExcelRow, columns: list[ExcelColumn]) -> None:
        """
        Classify a row as division, grouping, or item.

        Args:
            row: The row to classify
            columns: Detected columns
        """
        values = row.raw_values
        non_empty = [v for v in values if v is not None and str(v).strip()]

        # Check for division/section headers (usually bold, merged, or few columns filled)
        if len(non_empty) <= 2:
            text = " ".join(str(v) for v in non_empty).upper()
            # Common division keywords
            if any(kw in text for kw in ["DIVISION", "SECTION", "WORK", "PART"]):
                row.is_division = True
                return
            # All caps might indicate division
            if text == text.upper() and len(text) > 5:
                row.is_division = True
                return

        # Check for grouping (sub-section)
        if len(non_empty) <= 3:
            text = " ".join(str(v) for v in non_empty)
            # Lettered or numbered sub-sections
            if re.match(r"^[A-Za-z][\.\)]\s+\w+", text) or re.match(r"^\d+\.\d+\s+\w+", text):
                row.is_grouping = True
                return

        # Check for item (has quantity, unit, etc.)
        has_quantity = False
        has_description = False

        for col in columns:
            if col.index < len(values):
                val = values[col.index]
                if col.column_type == ColumnType.QUANTITY and val is not None:
                    try:
                        float(val)
                        has_quantity = True
                    except (ValueError, TypeError):
                        pass
                elif col.column_type == ColumnType.DESCRIPTION and val:
                    has_description = True

        if has_quantity or (has_description and len(non_empty) >= 3):
            row.is_item = True

    def _sheet_to_text(self, sheet: ExcelSheet) -> str:
        """
        Convert sheet to text representation for LLM processing.

        Args:
            sheet: Extracted sheet data

        Returns:
            Text representation
        """
        lines = []

        # Add column headers if detected
        if sheet.columns:
            headers = [col.header for col in sheet.columns if col.header]
            if headers:
                lines.append("COLUMNS: " + " | ".join(headers))
                lines.append("-" * 50)

        # Add rows
        for row in sheet.rows:
            if row.is_header:
                continue

            # Format row based on type
            values = [str(v) if v is not None else "" for v in row.raw_values]
            row_text = " | ".join(v for v in values if v.strip())

            if row.is_division:
                lines.append(f"\n[DIVISION] {row_text}")
            elif row.is_grouping:
                lines.append(f"\n  [GROUP] {row_text}")
            elif row.is_item:
                lines.append(f"    [ITEM] {row_text}")
            elif row_text.strip():
                lines.append(f"    {row_text}")

        return "\n".join(lines)


def extract_from_excel(file_path: str | Path) -> str:
    """
    Convenience function to extract text from Excel file.

    Args:
        file_path: Path to Excel file

    Returns:
        Extracted text
    """
    parser = ExcelParser(file_path)
    result = parser.extract()
    return result.text


def extract_excel_result(file_path: str | Path) -> ExcelExtractionResult:
    """
    Extract full result from Excel file.

    Args:
        file_path: Path to Excel file

    Returns:
        ExcelExtractionResult with full data
    """
    parser = ExcelParser(file_path)
    return parser.extract()


# TODO: Add support for .xls files (requires xlrd)
# TODO: Add merged cell handling
# TODO: Add support for multiple BOQ tables in one sheet
# TODO: Add currency detection from cell formatting
# TODO: Add support for formulas and cell references
